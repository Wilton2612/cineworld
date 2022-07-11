from openpyxl import Workbook
import openpyxl

from wtforms import validators
  



def validate_caracter_special_and_number(form,field):
    unido = field.data.replace(" ", "")
    if unido.isalpha() == False:
        raise validators.ValidationError('No se permiten caracteres especiales o números')

def validate_telefono(form, field):
    if field.data.isdigit() == False:
        raise validators.ValidationError('No se permiten caracteres')

def validate_choose(form, field):
    if field.data == '0':
        raise validators.ValidationError('Elige una opción valida')



def escribir(iden):
    try:
        hoja = Workbook()
        sheet = hoja.active
        sheet['A1'] = iden      
        hoja.save('src/data/dato.xlsx')
        return 0
    except Exception as e:
        print(e)
        return -1
       

def leer():
    book = openpyxl.load_workbook('src/data/dato.xlsx')
    libro = book.active
    usuario = libro['A1']
    return usuario.value

def delete():
    try:
        hoja = Workbook()
        sheet = hoja.active
        sheet.delete_cols(1)     
        hoja.save('src/data/dato.xlsx')
        return 0
    except Exception as e:
        print(e, "Intento fallido")
        return -1

